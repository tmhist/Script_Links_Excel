import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import random
import urllib.parse
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# User agents so that the search engine does not think that the search is performed by a bot
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)",
    "Mozilla/5.0 (X11; Linux x86_64)",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 15_2 like Mac OS X)",
    "Mozilla/5.0 (iPad; CPU OS 14_0 like Mac OS X)"
]


# Extract real URL from DuckDuckGo redirect link
def extract_real_url(duckduckgo_url):
    parsed = urllib.parse.urlparse(duckduckgo_url)
    query = urllib.parse.parse_qs(parsed.query)
    if "uddg" in query:
        return urllib.parse.unquote(query["uddg"][0])
    return duckduckgo_url


# Search for website via DuckDuckGo
def get_website(organisation_name):
    try:
        query = organisation_name.strip()
        url = f"https://html.duckduckgo.com/html?q={requests.utils.quote(query)}"
        headers = {"User-Agent": random.choice(USER_AGENTS)}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")
        result = soup.find("a", class_="result__a", href=True)
        if result:
            redirect_url = result["href"]
            real_url = extract_real_url(redirect_url)
            return real_url
    except Exception as e:
        print(f"Fehler bei der Suche nach Website für '{organisation_name}': {e}")
    return None


# Adds hyperlinks only
def update_excel_with_links(input_file, output_file, sheet_name="ZER-GmbH-20250612"):
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    # Readd column headings
    headers = [cell.value for cell in ws[1]]
    if "Organisation" not in headers:
        raise ValueError("Spalte 'Organisation' fehlt im Excel-Blatt")
    
    # Column index
    org_col = headers.index("Organisation") + 1
    if "Webseite" not in headers:
        ws.cell(row=1, column=len(headers)+1).value = "Webseite"
        web_col = len(headers)+1
    else:
        web_col = headers.index("Webseite") + 1

    # Iterate through every organization
    for row in range(2, ws.max_row + 1):
        org_name = ws.cell(row=row, column=org_col).value
        if not org_name:
            continue

        print(f"Suche Webseite für: {org_name}")
        site = get_website(org_name)
        if site:
            print(f"Gefunden: {site}\n")
            cell = ws.cell(row=row, column=web_col)
            cell.value = site
            cell.hyperlink = site
            cell.font = Font(color="0000EE", underline="single")
        else:
            print("Keine Webseite gefunden.\n")
        time.sleep(random.uniform(5, 9))

    wb.save(output_file)
    print(f"Gespeichert: {output_file}")


# Start
update_excel_with_links("ZER-GmbH-20250612.xlsx", "ZER-GmbH-20250612_Links.xlsx")